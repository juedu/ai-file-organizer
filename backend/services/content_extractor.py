"""문서 텍스트 추출 서비스"""

import logging
from pathlib import Path

logger = logging.getLogger(__name__)

MAX_CONTENT_LENGTH = 1000


class ContentExtractor:
    def extract(self, file_path: Path) -> str | None:
        """파일에서 텍스트를 추출합니다. 지원하지 않는 형식이면 None 반환."""
        ext = file_path.suffix.lower()

        extractors = {
            ".pdf": self._extract_pdf,
            ".docx": self._extract_docx,
            ".xlsx": self._extract_xlsx,
            ".txt": self._extract_text,
            ".md": self._extract_text,
            ".csv": self._extract_text,
            ".rtf": self._extract_text,
        }

        extractor = extractors.get(ext)
        if not extractor:
            return None

        try:
            content = extractor(file_path)
            if content:
                return content[:MAX_CONTENT_LENGTH]
            return None
        except Exception as e:
            logger.warning("텍스트 추출 실패 %s: %s", file_path.name, e)
            return None

    def _extract_pdf(self, file_path: Path) -> str | None:
        try:
            from PyPDF2 import PdfReader
            reader = PdfReader(str(file_path))
            texts = []
            for page in reader.pages[:5]:  # 최대 5페이지
                text = page.extract_text()
                if text:
                    texts.append(text)
            return "\n".join(texts) if texts else None
        except Exception as e:
            logger.warning("PDF 추출 실패 %s: %s", file_path.name, e)
            return None

    def _extract_docx(self, file_path: Path) -> str | None:
        try:
            from docx import Document
            doc = Document(str(file_path))
            texts = [p.text for p in doc.paragraphs if p.text.strip()]
            return "\n".join(texts) if texts else None
        except Exception as e:
            logger.warning("DOCX 추출 실패 %s: %s", file_path.name, e)
            return None

    def _extract_xlsx(self, file_path: Path) -> str | None:
        try:
            from openpyxl import load_workbook
            wb = load_workbook(str(file_path), read_only=True, data_only=True)
            texts = []
            for sheet in wb.worksheets[:3]:  # 최대 3시트
                for row in sheet.iter_rows(max_row=20, values_only=True):
                    row_text = " ".join(str(c) for c in row if c is not None)
                    if row_text.strip():
                        texts.append(row_text)
            wb.close()
            return "\n".join(texts) if texts else None
        except Exception as e:
            logger.warning("XLSX 추출 실패 %s: %s", file_path.name, e)
            return None

    def _extract_text(self, file_path: Path) -> str | None:
        for encoding in ("utf-8", "cp949", "euc-kr", "latin-1"):
            try:
                text = file_path.read_text(encoding=encoding)
                return text if text.strip() else None
            except (UnicodeDecodeError, ValueError):
                continue
        return None
